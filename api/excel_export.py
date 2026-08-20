"""
Excel export — panelde gösterilen veriden .xlsx üretir. Vercel serverless
uyumlu: dosya diske yazılmaz, her şey BytesIO üzerinde bellekte kalır.
"""
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, LineChart, Reference

ARIAL = "Arial"
H_FONT = Font(name=ARIAL, bold=True, size=13, color="FFFFFF")
SEC_FONT = Font(name=ARIAL, bold=True, size=10, color="FFFFFF")
LBL = Font(name=ARIAL, size=10)
LBLB = Font(name=ARIAL, bold=True, size=10)
NOTE = Font(name=ARIAL, italic=True, size=9, color="808080")
NAVY = PatternFill("solid", fgColor="1F3B57")
BLUE = PatternFill("solid", fgColor="2E5F8A")
TEAL = PatternFill("solid", fgColor="3C7A89")
LIGHT = PatternFill("solid", fgColor="EAF1F8")
GREEN = PatternFill("solid", fgColor="D9EAD3")
AMBER = PatternFill("solid", fgColor="FCE5CD")
RED = PatternFill("solid", fgColor="F4CCCC")
GREY = PatternFill("solid", fgColor="F2F2F2")
thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _section(ws, row, col_span, title, fill=BLUE):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_span)
    c = ws.cell(row, 1, title)
    c.font = SEC_FONT
    c.fill = fill
    c.alignment = LEFT
    return row + 1


def _kv_rows(ws, row, pairs, col_a=1, col_b=3, span=4):
    for label, val in pairs:
        ws.merge_cells(start_row=row, start_column=col_a, end_row=row, end_column=col_a + 1)
        lc = ws.cell(row, col_a, label)
        lc.font = LBL
        lc.alignment = LEFT
        lc.fill = LIGHT
        lc.border = BORDER
        ws.cell(row, col_a + 1).border = BORDER
        ws.cell(row, col_a + 1).fill = LIGHT
        ws.merge_cells(start_row=row, start_column=col_b, end_row=row, end_column=col_a + span - 1)
        vc = ws.cell(row, col_b, val if val is not None else "n/a")
        vc.font = LBLB
        vc.alignment = LEFT
        vc.border = BORDER
        for c in range(col_b + 1, col_a + span):
            ws.cell(row, c).border = BORDER
        row += 1
    return row


def build_report_xlsx(data: dict) -> bytes:
    """
    Panelin /api/analyze yanıtından (frontend'in gösterdiği `data` objesi)
    tek sayfalık kapsamlı bir Excel raporu üretir: pazar kararı + ön
    değerlendirme + pazar özeti + grafikler + top rakipler + relevant
    keywords + kar analizi girdi alanları.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Analiz Raporu"[:31]
    ws.sheet_view.showGridLines = False
    widths = {"A": 30, "B": 14, "C": 16, "D": 14, "E": 12, "F": 12, "G": 12, "H": 12, "I": 14, "J": 10}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    keyword = data.get("keyword", "")
    marketplace = data.get("marketplace", "")
    category_used = data.get("category_used", "")

    ws.merge_cells("A1:J1")
    c = ws.cell(1, 1, f"PL PAZAR ANALİZ RAPORU — “{keyword}” | {marketplace}")
    c.font = H_FONT
    c.fill = NAVY
    c.alignment = CENTER
    ws.row_dimensions[1].height = 26

    ws.merge_cells("A2:J2")
    ws.cell(2, 1, f"Kategori: {category_used or 'n/a'}").font = NOTE

    row = 4
    pa = data.get("pre_assessment", {})
    row = _section(ws, row, 10, f"PAZAR KARARI (ön öneri): {pa.get('verdict', 'n/a')}  ·  Olumsuz kriter: {pa.get('negative_count', 'n/a')}/6")
    row += 1

    row = _section(ws, row, 10, "ÖN DEĞERLENDİRME")
    crit_start = row
    for crit in pa.get("criteria", []):
        ws.cell(row, 1, crit["label"]).font = LBL
        ws.cell(row, 1).border = BORDER
        val = crit.get("value")
        # Panelle AYNI hata Excel'de de vardı: büyüklüğe bakıp tahmin ediyordu,
        # "Güçlü Yeni Marka: 2" -> "%200,0" çıkıyordu. Artık backend'in
        # gönderdiği unit alanı kullanılıyor (percent / usd / count).
        unit = crit.get("unit", "percent")
        if val is None:
            val_display = "n/a"
        elif not isinstance(val, (int, float)):
            val_display = val
        elif unit == "count":
            val_display = str(val)
        elif unit == "usd":
            val_display = f"${val:,.2f}"
        else:
            val_display = f"{val*100:.1f}%"
        ws.cell(row, 2, val_display).font = LBLB
        ws.cell(row, 2).border = BORDER
        flag_cell = ws.cell(row, 3, crit.get("flag", ""))
        flag_cell.border = BORDER
        flag_cell.font = Font(name=ARIAL, bold=True, size=10)
        if crit.get("flag") == "OK":
            flag_cell.fill = GREEN
        elif crit.get("flag") == "OLUMSUZ":
            flag_cell.fill = RED
        else:
            flag_cell.fill = GREY
        row += 1

    row += 1
    row = _section(ws, row, 10, "PAZAR ÖZETİ")
    stats = data.get("market_stats", {}) or {}
    stat_pairs = [
        ("Ort. Fiyat", f"${stats.get('avgPrice', 'n/a')}" if stats.get("avgPrice") else "n/a"),
        ("Ort. Rating", stats.get("avgRating", "n/a")),
        ("Ort. Review", stats.get("avgRatings", "n/a")),
        ("Toplam Marka", stats.get("brands", "n/a")),
        ("Ort. Satıcı", stats.get("avgSellers", "n/a")),
        ("Yeni Ürün (12ay)", stats.get("newProducts", "n/a")),
        ("İlk Listing", stats.get("firstShelfDate", "n/a")),
    ]
    row = _kv_rows(ws, row, stat_pairs, span=4)

    # --- Marka payı grafiği ---
    row += 1
    brand_items = data.get("brand_concentration", []) or []
    if brand_items:
        chart_start = row
        ws.cell(row, 1, "Marka").font = SEC_FONT
        ws.cell(row, 1).fill = TEAL
        ws.cell(row, 2, "Ciro Payı").font = SEC_FONT
        ws.cell(row, 2).fill = TEAL
        row += 1
        first_data_row = row
        for b in brand_items[:10]:
            ws.cell(row, 1, b.get("brand", "?")).border = BORDER
            share = b.get("totalRevenueRatio", 0)
            share = share if share and share <= 1 else (share / 100 if share else 0)
            ws.cell(row, 2, share).number_format = "0.0%"
            ws.cell(row, 2).border = BORDER
            row += 1
        last_data_row = row - 1
        pie = PieChart()
        pie.title = "Marka Payı"
        pie.height, pie.width = 7, 10
        data_ref = Reference(ws, min_col=2, min_row=first_data_row, max_row=last_data_row)
        cats_ref = Reference(ws, min_col=1, min_row=first_data_row, max_row=last_data_row)
        pie.add_data(data_ref)
        pie.set_categories(cats_ref)
        ws.add_chart(pie, f"D{chart_start}")
        row = max(row, chart_start + 15) + 1

    # --- Top Rakipler ---
    row += 1
    row = _section(ws, row, 10, "TOP RAKİPLER (otomatik çekildi)")
    headers = ["ASIN", "Marka", "Fiyat", "Aylık Satış", "Aylık Ciro", "BSR", "Rating", "Review", "Fulfillment"]
    for j, h in enumerate(headers, 1):
        c = ws.cell(row, j, h)
        c.font = SEC_FONT
        c.fill = BLUE
        c.alignment = CENTER
        c.border = BORDER
    row += 1
    for comp in data.get("top_competitors", []) or []:
        vals = [comp.get("asin"), comp.get("brand"), comp.get("price"), comp.get("units"),
                comp.get("revenue"), comp.get("bsr"), comp.get("rating"), comp.get("ratings"),
                comp.get("fulfillment")]
        for j, v in enumerate(vals, 1):
            cell = ws.cell(row, j, v if v is not None else "n/a")
            cell.border = BORDER
            cell.font = LBL
        row += 1

    # --- Relevant Keywords ---
    row += 1
    row = _section(ws, row, 10, "RELEVANT KEYWORDS (CVR/ACOS hesaplanan)")
    kw_headers = ["Keyword", "Search", "Clicks", "Purchases", "Click CVR", "Bid", "ACOS", "CPA", "Relevancy"]
    for j, h in enumerate(kw_headers, 1):
        c = ws.cell(row, j, h)
        c.font = SEC_FONT
        c.fill = BLUE
        c.alignment = CENTER
        c.border = BORDER
    row += 1
    for kw in data.get("keyword_rows", []) or []:
        vals = [kw.get("keyword"), kw.get("searches"), kw.get("clicks"), kw.get("purchases"),
                kw.get("click_cvr"), kw.get("bid"), kw.get("acos"), kw.get("cpa"), kw.get("relevancy")]
        for j, v in enumerate(vals, 1):
            cell = ws.cell(row, j, v if v is not None else "n/a")
            cell.border = BORDER
            cell.font = LBL
            if j in (5, 7) and isinstance(v, (int, float)):
                cell.number_format = "0.0%"
            if j == 6 and isinstance(v, (int, float)):
                cell.number_format = "$#,##0.00"
        row += 1

    # --- Kar Analizi (canlı formüllerle — panelde girilen değerlerle önceden doldurulur) ---
    row += 1
    row = _section(ws, row, 10, "KAR ANALİZİ  (sarı hücreler manuel — değiştirince otomatik yeniden hesaplanır)")
    profit = data.get("profit_analysis") or {}
    inputs = profit.get("inputs", {})
    YELLOW = PatternFill("solid", fgColor="FFF2CC")
    INPUT_FONT = Font(name=ARIAL, size=10, color="0000FF")
    FORMULA_FONT = Font(name=ARIAL, size=10, color="000000")

    def _plabel(r, text, bold=False):
        c = ws.cell(r, 1, text)
        c.font = LBLB if bold else LBL
        c.border = BORDER
        ws.cell(r, 2).border = BORDER
        return r + 1

    def _pinput(r, val, fmt):
        cell = ws.cell(r, 2, val if val is not None else 0)
        cell.font = INPUT_FONT
        cell.fill = YELLOW
        cell.border = BORDER
        cell.number_format = fmt
        ws.cell(r, 1).border = BORDER

    def _pformula(r, formula, fmt, bold=False):
        cell = ws.cell(r, 2, formula)
        cell.font = Font(name=ARIAL, bold=bold, size=10, color="000000")
        cell.border = BORDER
        cell.number_format = fmt
        ws.cell(r, 1).border = BORDER

    r_cogs = row; row = _plabel(row, "Alış Fiyatı / COGS ($)"); _pinput(r_cogs, inputs.get("cogs"), '$#,##0.00')
    r_sale = row; row = _plabel(row, "Satış Fiyatı ($)"); _pinput(r_sale, inputs.get("sale"), '$#,##0.00')
    r_fba = row; row = _plabel(row, "FBA Fee ($)"); _pinput(r_fba, inputs.get("fba"), '$#,##0.00')
    r_ref = row; row = _plabel(row, "Referral Fee ($)"); _pinput(r_ref, inputs.get("ref"), '$#,##0.00')
    r_acos = row; row = _plabel(row, "ACOS (%)"); _pinput(r_acos, (inputs.get("acos") or 0) / 100, '0.0%')
    r_ret = row; row = _plabel(row, "Return Rate (%)"); _pinput(r_ret, (inputs.get("ret") or 0) / 100, '0.0%')
    r_gen = row; row = _plabel(row, "Genel Gider (%)"); _pinput(r_gen, (inputs.get("gen") or 0) / 100, '0.0%')

    r_adv = row; row = _plabel(row, "Reklam Maliyeti ($)"); _pformula(r_adv, f"=B{r_acos}*B{r_sale}", '$#,##0.00')
    r_genc = row; row = _plabel(row, "Genel Gider ($)"); _pformula(r_genc, f"=B{r_gen}*B{r_sale}", '$#,##0.00')
    r_retc = row; row = _plabel(row, "Return Maliyeti ($)"); _pformula(r_retc, f"=B{r_ret}*(B{r_cogs}+B{r_fba})", '$#,##0.00')
    r_tot = row; row = _plabel(row, "Toplam Maliyet ($)"); _pformula(r_tot, f"=B{r_cogs}+B{r_fba}+B{r_ref}+B{r_adv}+B{r_genc}+B{r_retc}", '$#,##0.00')
    r_profit = row; row = _plabel(row, "Birim Kar / Zarar ($)", bold=True); _pformula(r_profit, f"=B{r_sale}-B{r_tot}", '$#,##0.00;[Red]($#,##0.00)', bold=True)
    r_margin = row; row = _plabel(row, "Kar Marjı (%)", bold=True); _pformula(r_margin, f"=IF(B{r_sale}=0,0,B{r_profit}/B{r_sale})", '0.0%;[Red](0.0%)', bold=True)
    r_roi = row; row = _plabel(row, "ROI (%) = Kar/COGS", bold=True); _pformula(r_roi, f"=IF(B{r_cogs}=0,0,B{r_profit}/B{r_cogs})", '0.0%;[Red](0.0%)', bold=True)

    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    ws.cell(row, 1, "Bu rapor PL Pazar Paneli tarafından canlı SellerSprite MCP verisiyle otomatik üretilmiştir.").font = NOTE

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def build_keywords_xlsx(keyword_rows: list, seed_keyword: str = "") -> bytes:
    """Sadece Relevant Keywords tablosunu ayrı bir Excel dosyası olarak üretir."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Keywords"
    ws.sheet_view.showGridLines = False
    for col, w in {"A": 34, "B": 12, "C": 10, "D": 11, "E": 11, "F": 9, "G": 9, "H": 10, "I": 10}.items():
        ws.column_dimensions[col].width = w

    ws.merge_cells("A1:I1")
    c = ws.cell(1, 1, f"RELEVANT KEYWORDS — “{seed_keyword}”")
    c.font = H_FONT
    c.fill = NAVY
    c.alignment = CENTER
    ws.row_dimensions[1].height = 24

    headers = ["Keyword", "Search", "Clicks", "Purchases", "Click CVR", "Bid", "ACOS", "CPA", "Relevancy"]
    for j, h in enumerate(headers, 1):
        cell = ws.cell(3, j, h)
        cell.font = SEC_FONT
        cell.fill = BLUE
        cell.alignment = CENTER
        cell.border = BORDER
    row = 4
    for kw in keyword_rows:
        vals = [kw.get("keyword"), kw.get("searches"), kw.get("clicks"), kw.get("purchases"),
                kw.get("click_cvr"), kw.get("bid"), kw.get("acos"), kw.get("cpa"), kw.get("relevancy")]
        for j, v in enumerate(vals, 1):
            cell = ws.cell(row, j, v if v is not None else "n/a")
            cell.border = BORDER
            cell.font = LBL
            if j in (5, 7) and isinstance(v, (int, float)):
                cell.number_format = "0.0%"
            if j == 6 and isinstance(v, (int, float)):
                cell.number_format = "$#,##0.00"
        if (row - 4) % 2 == 1:
            for j in range(1, 10):
                ws.cell(row, j).fill = GREY
        row += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
